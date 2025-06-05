import pandas as pd
import numpy as np
from PIL import Image
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import io

# -------------------------------
# RAW画像仕様（必要に応じて変更）
# -------------------------------
width = 224       # 画像幅
height = 224      # 画像高さ
channels = 1      # グレースケール=1、RGB=3

# -------------------------------
# Excelファイルの読み込み
# -------------------------------
input_path = "input.xlsx"
q_sheet = pd.read_excel(input_path, sheet_name="Q")
b_sheet = pd.read_excel(input_path, sheet_name="B")

# -------------------------------
# Qシート290行目から処理
# -------------------------------
results = []

import pandas as pd

# 全体読み込み
df_all = pd.read_excel("input.xlsx", sheet_name="Q")

# A列に "confused pairs" がある行を検索
target_rows = df_all[df_all.iloc[:, 0] == 'confused pairs']

if not target_rows.empty:
    confused_start_idx = target_rows.index[0] + 1
    df_confused = df_all.iloc[confused_start_idx:].reset_index(drop=True)
    print(df_confused.head())
else:
    print("❌ 'confused pairs' 行が見つかりませんでした。")


for idx in range(289, len(q_sheet)):  # Excelの290行目から
    entry = q_sheet.loc[idx, 'A']
    if not isinstance(entry, str) or "_" not in entry:
        continue

    parts = entry.split('_')
    if len(parts) < 4:
        continue

    true_class = f"{parts[0]}_{parts[1]}"
    pred_class = f"{parts[2]}_{parts[3]}"

    matches = b_sheet[(b_sheet['True'] == true_class) & (b_sheet['Pred'] == pred_class)]

    for _, row in matches.iterrows():
        raw_file = row['RawFile']
        image_path = os.path.join("Wrong images", true_class, raw_file)
        results.append({
            'index': len(results),
            'classname': true_class,
            'imagename': raw_file,
            'image_path': image_path
        })

# -------------------------------
# Excelのワークブックに result シート追加
# -------------------------------
wb = load_workbook(input_path)
if 'result' in wb.sheetnames:
    del wb['result']
ws = wb.create_sheet(title='result')
ws.append(['index', 'classname', 'imagename', 'image'])

# -------------------------------
# 結果を1行ずつ貼り付け＋画像処理
# -------------------------------
for row in results:
    r = ws.max_row + 1
    ws.cell(row=r, column=1, value=row['index'])
    ws.cell(row=r, column=2, value=row['classname'])
    ws.cell(row=r, column=3, value=row['imagename'])

    raw_path = row['image_path']
    if os.path.exists(raw_path):
        with open(raw_path, 'rb') as f:
            raw_data = f.read()
            img_array = np.frombuffer(raw_data, dtype=np.uint8)
            if len(img_array) == width * height * channels:
                if channels == 1:
                    img_array = img_array.reshape((height, width))
                    pil_img = Image.fromarray(img_array, 'L')
                else:
                    img_array = img_array.reshape((height, width, channels))
                    pil_img = Image.fromarray(img_array, 'RGB')

                img_io = io.BytesIO()
                pil_img.save(img_io, format='PNG')
                img_io.seek(0)
                xl_img = XLImage(img_io)
                ws.add_image(xl_img, f'D{r}')

# -------------------------------
# 上書き保存
# -------------------------------
wb.save(input_path)
print("✅ 完了: 結果を input.xlsx の result シートに保存しました。")


import pandas as pd
import numpy as np
from PIL import Image
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
import os
import io

# RAW画像の仕様
width = 224
height = 224
channels = 1  # グレースケール

# Excelファイルパス
input_path = "input.xlsx"

# QシートとBシートの読み込み
df_all = pd.read_excel(input_path, sheet_name="Q")
df_b = pd.read_excel(input_path, sheet_name="B")

# "confused pairs" 行を探す
confused_rows = df_all[df_all.iloc[:, 0] == "confused pairs"]
if confused_rows.empty:
    print("❌ 'confused pairs' 行が見つかりませんでした。")
    exit()

start_idx = confused_rows.index[0] + 1
df_q = df_all.iloc[start_idx:].reset_index(drop=True)

# 結果リスト
results = []

# データ走査
for idx, row in df_q.iterrows():
    entry = row.iloc[0]
    if not isinstance(entry, str) or "_" not in entry:
        continue

    parts = entry.split('_')
    if len(parts) < 4:
        continue

    true_class = f"{parts[0]}_{parts[1]}"
    pred_class = f"{parts[2]}_{parts[3]}"

    matches = df_b[(df_b['True'] == true_class) & (df_b['Pred'] == pred_class)]
    for _, m in matches.iterrows():
        image_path = os.path.join("Wrong images", true_class, m['RawFile'])
        results.append({
            'index': len(results),
            'classname': true_class,
            'imagename': m['RawFile'],
            'image_path': image_path
        })

# ワークブックに追記
wb = load_workbook(input_path)
if 'result' in wb.sheetnames:
    del wb['result']
ws = wb.create_sheet("result")
ws.append(['index', 'classname', 'imagename', 'image'])

# 画像付きで書き出し
for row in results:
    r = ws.max_row + 1
    ws.cell(row=r, column=1, value=row['index'])
    ws.cell(row=r, column=2, value=row['classname'])
    ws.cell(row=r, column=3, value=row['imagename'])

    raw_path = row['image_path']
    if os.path.exists(raw_path):
        with open(raw_path, 'rb') as f:
            raw_data = f.read()
            img_array = np.frombuffer(raw_data, dtype=np.uint8)
            if len(img_array) == width * height * channels:
                if channels == 1:
                    img_array = img_array.reshape((height, width))
                    pil_img = Image.fromarray(img_array, 'L')
                else:
                    img_array = img_array.reshape((height, width, channels))
                    pil_img = Image.fromarray(img_array, 'RGB')
                img_io = io.BytesIO()
                pil_img.save(img_io, format='PNG')
                img_io.seek(0)
                xl_img = XLImage(img_io)
                ws.add_image(xl_img, f'D{r}')

# 保存
wb.save(input_path)
print("✅ 完了: 結果を 'result' シートに保存しました。")


