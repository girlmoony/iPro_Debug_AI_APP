import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage

# Excelファイル読み込み
df = pd.read_excel('classification_scores.xlsx')

# 画像保存とExcel貼り付け用の関数
def plot_and_save(x_col, y_col, label_col, title, img_filename):
    # 散布図作成
    plt.figure(figsize=(6, 6))
    plt.scatter(df[x_col], df[y_col])
    plt.plot([0, 1], [0, 1], 'r--')
    for i in range(len(df)):
        plt.text(df[x_col][i] + 0.005, df[y_col][i], str(df[label_col][i]))
    plt.xlabel(x_col)
    plt.ylabel(y_col)
    plt.title(title)
    plt.grid(True)
    plt.tight_layout()
    
    # 画像として保存
    plt.savefig(img_filename)
    plt.close()

# 3つの図を保存
plot_and_save('ipro_f1', 'gpu_f1', 'class', 'F1 Score Comparison', 'f1.png')
plot_and_save('ipro_precision', 'gpu_precision', 'class', 'Precision Comparison', 'precision.png')
plot_and_save('ipro_recall', 'gpu_recall', 'class', 'Recall Comparison', 'recall.png')

# Excelファイルに画像を貼り付ける
wb = load_workbook('classification_scores.xlsx')
if 'Graphs' not in wb.sheetnames:
    ws = wb.create_sheet('Graphs')
else:
    ws = wb['Graphs']

# 各画像をExcelシートに貼り付け
img_files = ['f1.png', 'precision.png', 'recall.png']
positions = ['A1', 'A20', 'A40']  # 位置調整

for img_file, pos in zip(img_files, positions):
    img = ExcelImage(img_file)
    ws.add_image(img, pos)

# 保存
wb.save('classification_scores.xlsx')
