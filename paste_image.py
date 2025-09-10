from pathlib import Path
from PIL import Image as PILImage
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

# ===== 設定 =====
excel_path = "input.xlsx"         # 読み込むExcel
sheet_name = "Sheet1"             # シート名
save_path = "output_with_images.xlsx"  # 出力ファイル
base_path = Path(r"/path/to/images_root")  # 画像の親フォルダ

# サムネイル最大サイズ（ピクセル）
MAX_W_PX = 120
MAX_H_PX = 120

# 列幅（B列）を少し広めに（Excel幅単位）
B_COL_WIDTH = 25  # 目安：25 ≒ 約180〜190px 相当

# ===== ユーティリティ =====
def ensure_png(name: str) -> str:
    n = (name or "").strip()
    if not n.lower().endswith(".png"):
        n += ".png"
    return n

def px_to_points(px: int, dpi: int = 96) -> float:
    # Excel行の高さはポイント指定。1pt = 1/72 inch、pxは一般的に96dpi前提で換算。
    return px * 72 / dpi

# ===== 実装 =====
wb = load_workbook(excel_path)
ws = wb[sheet_name]

# B列の幅を調整（画像が見切れないように）
ws.column_dimensions["B"].width = B_COL_WIDTH

# 見出し行が1行ある想定で2行目からループ（必要に応じて変更）
start_row = 2
max_row = ws.max_row

not_found = []

for r in range(start_row, max_row + 1):
    file_name_cell = ws.cell(row=r, column=2)  # B列：pngファイル名
    folder_cell    = ws.cell(row=r, column=3)  # C列：フォルダ名

    file_name = ensure_png(file_name_cell.value if file_name_cell.value is not None else "")
    folder_name = (folder_cell.value or "").strip()

    if not file_name or not folder_name:
        continue  # どちらか欠けている行はスキップ

    img_path = base_path / folder_name / file_name

    if not img_path.exists():
        not_found.append(str(img_path))
        continue

    # 画像を開いてサムネイル化（縦横比維持）
    with PILImage.open(img_path) as im:
        w, h = im.size
        scale = min(MAX_W_PX / w, MAX_H_PX / h, 1.0)
        new_w = int(w * scale)
        new_h = int(h * scale)

    ximg = XLImage(str(img_path))
    # openpyxlのImageはwidth/height（px相当）を直接セット可能
    ximg.width = new_w
    ximg.height = new_h

    # B列セルにアンカー（画像は「配置」されるだけでテキストは残せます）
    # ファイル名が邪魔ならBセルは空にしてOK
    file_name_cell.value = None
    ximg.anchor = f"B{r}"
    ws.add_image(ximg)

    # 行高を画像高さに合わせて調整
    ws.row_dimensions[r].height = px_to_points(new_h)

# 保存
wb.save(save_path)

# 見つからなかったパスのログ（任意）
if not_found:
    print("[NOT FOUND]")
    for p in not_found:
        print(" -", p)

print(f"書き出し完了: {save_path}")
